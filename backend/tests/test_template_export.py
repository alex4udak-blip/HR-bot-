"""Тесты генерации шаблона/экспорта сотрудников."""
import io
from datetime import datetime
from types import SimpleNamespace

import openpyxl


def _emp(**kw):
    base = dict(user=SimpleNamespace(email='ivanov@mst.io'), position='Разработчик',
               phone='+7 (999) 123-45-67', telegram_username='@ivanov',
               department_start_date=datetime(2025, 1, 15),
               extra_data={'last_name': 'Иванов', 'first_name': 'Иван', 'middle_name': 'Ив.',
                           'address': 'Москва', 'passport_number': '4509 123456',
                           'payment_method': 'card', 'payment_details': '2200...'})
    base.update(kw)
    return SimpleNamespace(**base)


def test_normalize_pay_method_label_to_key():
    from api.routes.employees import _normalize_pay_method
    assert _normalize_pay_method('Карта') == 'card'
    assert _normalize_pay_method('Криптокошелёк') == 'crypto'
    assert _normalize_pay_method('card') == 'card'
    assert _normalize_pay_method('нечто') == 'нечто'
    assert _normalize_pay_method(None) is None


def test_employee_to_template_row_maps_fields():
    from api.routes.employees import _employee_to_template_row
    row = _employee_to_template_row(_emp())
    assert row['Email'] == 'ivanov@mst.io'
    assert row['Фамилия'] == 'Иванов'
    assert row['Должность'] == 'Разработчик'
    assert row['Дата начала (ДД.ММ.ГГГГ)'] == '15.01.2025'
    assert row['Способ выплаты'] == 'Карта'
    assert row['Реквизиты'] == '2200...'


def test_build_template_xlsx_blank_has_header_and_sheets():
    from api.routes.employees import _build_template_xlsx, TEMPLATE_HEADERS
    data = _build_template_xlsx(rows=[], positions=['Разработчик', 'Дизайнер'], with_example=True)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb['Сотрудники']
    assert [c.value for c in ws[1]] == TEMPLATE_HEADERS
    assert 'Инструкция' in wb.sheetnames
    assert ws['A2'].value not in (None, '')


def test_build_template_xlsx_filled_rows():
    from api.routes.employees import _build_template_xlsx, _employee_to_template_row
    rows = [_employee_to_template_row(_emp())]
    data = _build_template_xlsx(rows=rows, positions=['Разработчик'], with_example=False)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb['Сотрудники']
    assert ws['A2'].value == 'ivanov@mst.io'
