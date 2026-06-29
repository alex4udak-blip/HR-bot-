"""
Employee Personal Cabinet, Leave Counter, and Auto-Reminders routes.

Endpoints:
- CRUD for employees (admin/HRD)
- /me endpoint for employee self-service
- Leave balance & leave requests
- Auto-reminders for probation/anniversary dates
"""
import io
import logging
from datetime import datetime, timedelta
from typing import Optional, List

from dateutil.relativedelta import relativedelta
from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile, File, Response
from pydantic import BaseModel
from sqlalchemy import select, func, and_, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from api.database import get_db
from api.models.database import (
    Employee, LeaveRequest, User, Organization, Department,
    DepartmentMember, OrgUnit, OrgMember, OrgRole, Notification, EmployeeDocument,
)
from api.services.auth import get_current_user, get_user_org

logger = logging.getLogger("hr-analyzer.employees")

router = APIRouter()


# ─── openpyxl imports ──────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

TEMPLATE_HEADERS = ['Email', 'Фамилия', 'Имя', 'Отчество', 'Должность', 'Телефон', 'Telegram',
                    'Дата начала (ДД.ММ.ГГГГ)', 'Адрес', 'Паспорт №', 'Способ выплаты', 'Реквизиты']
_TEMPLATE_WIDTHS = [26, 16, 14, 16, 20, 20, 16, 24, 32, 16, 18, 30]
_PAY_LABELS = ['Карта', 'Криптокошелёк', 'Банковский счёт', 'Другое']
_PAY_KEY_TO_LABEL = {'card': 'Карта', 'crypto': 'Криптокошелёк', 'bank': 'Банковский счёт', 'other': 'Другое'}
_PAY_LABEL_TO_KEY = {v: k for k, v in _PAY_KEY_TO_LABEL.items()}


def _normalize_pay_method(v):
    if v is None:
        return None
    s = str(v).strip()
    return _PAY_LABEL_TO_KEY.get(s, s)


def _employee_to_template_row(emp) -> dict:
    e = emp.extra_data or {}
    def s(x):
        return '' if x is None else str(x)
    email = emp.user.email if getattr(emp, 'user', None) else ''
    date_str = emp.department_start_date.strftime('%d.%m.%Y') if emp.department_start_date else ''
    method = e.get('payment_method')
    return {
        'Email': s(email),
        'Фамилия': s(e.get('last_name')),
        'Имя': s(e.get('first_name')),
        'Отчество': s(e.get('middle_name')),
        'Должность': s(emp.position),
        'Телефон': s(emp.phone),
        'Telegram': s(emp.telegram_username),
        'Дата начала (ДД.ММ.ГГГГ)': date_str,
        'Адрес': s(e.get('address')),
        'Паспорт №': s(e.get('passport_number')),
        'Способ выплаты': _PAY_KEY_TO_LABEL.get(method, s(method)),
        'Реквизиты': s(e.get('payment_details')),
    }


def _build_template_xlsx(rows: list, positions: list, with_example: bool) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Сотрудники'
    head_fill = PatternFill('solid', fgColor='2F5496')
    key_fill = PatternFill('solid', fgColor='C00000')
    head_font = Font(bold=True, color='FFFFFF', size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ex_font = Font(italic=True, color='909090')
    for i, h in enumerate(TEMPLATE_HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = key_fill if i == 1 else head_fill
        c.font = head_font
        c.alignment = center
        c.border = border
    for i, w in enumerate(_TEMPLATE_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'
    last_col = get_column_letter(len(TEMPLATE_HEADERS))
    ws.auto_filter.ref = f'A1:{last_col}1'

    dv_pay = DataValidation(type='list', formula1='"%s"' % ','.join(_PAY_LABELS), allow_blank=True)
    dv_pay.promptTitle = 'Способ выплаты'
    dv_pay.prompt = 'Выберите из списка'
    ws.add_data_validation(dv_pay)
    pay_col = get_column_letter(TEMPLATE_HEADERS.index('Способ выплаты') + 1)
    dv_pay.add(f'{pay_col}2:{pay_col}1000')

    clean_pos = [p for p in dict.fromkeys(positions) if p]
    if clean_pos:
        lst = wb.create_sheet('Списки')
        for i, p in enumerate(clean_pos, start=1):
            lst.cell(row=i, column=1, value=p)
        lst.sheet_state = 'hidden'
        dv_pos = DataValidation(type='list', formula1=f'Списки!$A$1:$A${len(clean_pos)}', allow_blank=True)
        dv_pos.promptTitle = 'Должность'
        dv_pos.prompt = 'Выберите из текущих должностей компании'
        ws.add_data_validation(dv_pos)
        pos_col = get_column_letter(TEMPLATE_HEADERS.index('Должность') + 1)
        dv_pos.add(f'{pos_col}2:{pos_col}1000')

    dv_date = DataValidation(type='date', operator='greaterThan', formula1='date(2000,1,1)', allow_blank=True)
    dv_date.promptTitle = 'Дата начала'
    dv_date.prompt = 'Формат ДД.ММ.ГГГГ'
    ws.add_data_validation(dv_date)
    date_col = get_column_letter(TEMPLATE_HEADERS.index('Дата начала (ДД.ММ.ГГГГ)') + 1)
    dv_date.add(f'{date_col}2:{date_col}1000')

    r = 2
    if with_example and not rows:
        example = ['ivanov@mail.ru', 'Иванов', 'Иван', 'Иванович', 'Разработчик',
                   '+7 (999) 123-45-67', '@ivanov', '08.06.2026', 'г. Москва, ул. Ленина 1',
                   '4509 123456', 'Карта', '2200 1234 5678 9010']
        for i, v in enumerate(example, start=1):
            cc = ws.cell(row=r, column=i, value=v)
            cc.font = ex_font
            cc.border = border
        r += 1
    for row in rows:
        for i, h in enumerate(TEMPLATE_HEADERS, start=1):
            ws.cell(row=r, column=i, value=row.get(h, ''))
        r += 1

    ins = wb.create_sheet('Инструкция')
    ins.column_dimensions['A'].width = 28
    ins.column_dimensions['B'].width = 72
    ins.append(['Колонка', 'Как заполнять'])
    for cell in (ins['A1'], ins['B1']):
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = head_fill
    for pair in [
        ('Email', 'КЛЮЧ сопоставления — НЕ менять.'),
        ('Фамилия / Имя / Отчество', 'ФИО сотрудника.'),
        ('Должность', 'Выберите из списка или впишите.'),
        ('Телефон', '+7 (999) 123-45-67.'),
        ('Telegram', '@username.'),
        ('Дата начала (ДД.ММ.ГГГГ)', 'Напр. 08.06.2026.'),
        ('Адрес', 'Адрес проживания.'),
        ('Паспорт №', 'Серия и номер.'),
        ('Способ выплаты', 'Карта / Криптокошелёк / Банковский счёт / Другое.'),
        ('Реквизиты', 'Номер карты / адрес кошелька / банковские реквизиты.'),
    ]:
        ins.append(list(pair))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Pydantic schemas ───────────────────────────────────────

class EmployeeCreate(BaseModel):
    user_id: int
    entity_id: Optional[int] = None
    department_id: Optional[int] = None
    position: Optional[str] = None
    phone: Optional[str] = None
    telegram_username: Optional[str] = None
    practice_start_date: Optional[datetime] = None
    department_start_date: Optional[datetime] = None
    nda_signed: bool = False
    contract_signed: bool = False
    extra_data: Optional[dict] = None


class EmployeeUpdate(BaseModel):
    department_id: Optional[int] = None
    position: Optional[str] = None
    phone: Optional[str] = None
    telegram_username: Optional[str] = None
    practice_start_date: Optional[datetime] = None
    department_start_date: Optional[datetime] = None
    nda_signed: Optional[bool] = None
    nda_signed_at: Optional[datetime] = None
    contract_signed: Optional[bool] = None
    contract_signed_at: Optional[datetime] = None
    vacation_days_total: Optional[int] = None
    vacation_days_used: Optional[int] = None
    sick_days_total: Optional[int] = None
    sick_days_used: Optional[int] = None
    family_leave_days_total: Optional[int] = None
    family_leave_days_used: Optional[int] = None
    extra_data: Optional[dict] = None


class EmployeeResponse(BaseModel):
    id: int
    user_id: int
    org_id: int
    entity_id: Optional[int] = None
    department_id: Optional[int] = None
    position: Optional[str] = None
    phone: Optional[str] = None
    telegram_username: Optional[str] = None
    practice_start_date: Optional[datetime] = None
    department_start_date: Optional[datetime] = None
    probation_end_date: Optional[datetime] = None
    one_year_date: Optional[datetime] = None
    vacation_days_total: int = 0
    vacation_days_used: int = 0
    sick_days_total: int = 10
    sick_days_used: int = 0
    family_leave_days_total: int = 3
    family_leave_days_used: int = 0
    nda_signed: bool = False
    nda_signed_at: Optional[datetime] = None
    contract_signed: bool = False
    contract_signed_at: Optional[datetime] = None
    is_active: bool = True
    dismissed_at: Optional[datetime] = None
    dismissal_reason: Optional[str] = None
    extra_data: Optional[dict] = None
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None
    # Joined fields
    user_name: Optional[str] = None
    user_email: Optional[str] = None
    department_name: Optional[str] = None

    model_config = {"from_attributes": True}


class LeaveBalanceResponse(BaseModel):
    vacation_total: int
    vacation_used: int
    vacation_remaining: int
    sick_total: int
    sick_used: int
    sick_remaining: int
    family_leave_total: int
    family_leave_used: int
    family_leave_remaining: int
    cycle_start: Optional[datetime] = None
    cycle_end: Optional[datetime] = None


class LeaveRequestCreate(BaseModel):
    type: str  # vacation, sick, family_leave, bereavement
    start_date: datetime
    end_date: datetime
    days: int
    reason: Optional[str] = None


class LeaveRequestResponse(BaseModel):
    id: int
    employee_id: int
    type: str
    start_date: datetime
    end_date: datetime
    days: int
    reason: Optional[str] = None
    status: str
    approved_by: Optional[int] = None
    approved_at: Optional[datetime] = None
    created_at: Optional[datetime] = None
    # Joined
    employee_name: Optional[str] = None

    model_config = {"from_attributes": True}


class ReminderItem(BaseModel):
    employee_id: int
    employee_name: str
    type: str  # probation_ending, one_year_anniversary
    date: datetime
    days_remaining: int


class BulkImportRow(BaseModel):
    email: str
    last_name: Optional[str] = None
    first_name: Optional[str] = None
    middle_name: Optional[str] = None
    position: Optional[str] = None
    phone: Optional[str] = None
    telegram_username: Optional[str] = None
    department_start_date: Optional[str] = None  # ISO или ДД.ММ.ГГГГ
    address: Optional[str] = None
    passport_number: Optional[str] = None
    payment_method: Optional[str] = None
    payment_details: Optional[str] = None


# ─── Helpers ─────────────────────────────────────────────────

def _auto_calculate_dates(emp: Employee):
    """Auto-calculate probation_end_date and one_year_date from department_start_date."""
    if emp.department_start_date:
        emp.probation_end_date = emp.department_start_date + relativedelta(months=3)
        emp.one_year_date = emp.department_start_date + relativedelta(years=1)
    else:
        emp.probation_end_date = None
        emp.one_year_date = None


def _calculate_vacation_days(
    department_start_date: Optional[datetime], now: Optional[datetime] = None
) -> int:
    """Накоплено отпуска в ТЕКУЩЕМ рабочем году: 2 дня за каждый полный
    отработанный месяц цикла; СБРОС В НОЛЬ ровно на годовщине приёма
    (use-it-or-lose-it — неиспользованные дни сгорают в годовщину).
    Прогрессия: 0,2,4,…,22, затем на годовщине снова 0.
    `now` инъектируется в тестах для детерминизма; по умолчанию — текущий UTC."""
    if not department_start_date:
        return 0
    if now is None:
        now = datetime.utcnow()
    if now < department_start_date:
        return 0
    delta = relativedelta(now, department_start_date)
    months = delta.years * 12 + delta.months  # полных месяцев с даты приёма
    months_in_cycle = months % 12  # 0..11 (на годовщине обнуляется)
    return months_in_cycle * 2  # 0..22


def _parse_import_date(s: str) -> Optional[datetime]:
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:10], fmt)
        except ValueError:
            continue
    return None


def _merge_extra(base: Optional[dict], patch: Optional[dict]) -> dict:
    """Слить patch поверх base, сохранив все прочие ключи base (passport, documents и т.п.)."""
    out = dict(base or {})
    for k, v in (patch or {}).items():
        out[k] = v
    return out


def _cycle_start(
    department_start_date: Optional[datetime], now: Optional[datetime] = None
) -> Optional[datetime]:
    """Дата начала текущего рабочего года (последняя годовщина приёма;
    сама годовщина уже относится к новому циклу).
    `now` инъектируется в тестах; по умолчанию — текущий UTC."""
    if not department_start_date:
        return None
    if now is None:
        now = datetime.utcnow()
    if now < department_start_date:
        # Приём ещё не наступил — цикл формально начинается с даты приёма.
        return department_start_date
    delta = relativedelta(now, department_start_date)
    months = delta.years * 12 + delta.months
    cycle_index = months // 12
    return department_start_date + relativedelta(months=cycle_index * 12)


async def _is_admin_or_owner(user: User, org_id: int, db: AsyncSession) -> bool:
    """Check if user is superadmin or org owner/admin."""
    if user.role and user.role.value == "superadmin":
        return True
    result = await db.execute(
        select(OrgMember).where(
            OrgMember.user_id == user.id,
            OrgMember.org_id == org_id,
            OrgMember.role.in_([OrgRole.owner, OrgRole.admin]),
        )
    )
    return result.scalar_one_or_none() is not None


def _employee_to_response(emp: Employee) -> EmployeeResponse:
    """Convert Employee ORM object to response, including joined fields."""
    return EmployeeResponse(
        id=emp.id,
        user_id=emp.user_id,
        org_id=emp.org_id,
        entity_id=emp.entity_id,
        department_id=emp.department_id,
        position=emp.position,
        phone=emp.phone,
        telegram_username=emp.telegram_username,
        practice_start_date=emp.practice_start_date,
        department_start_date=emp.department_start_date,
        probation_end_date=emp.probation_end_date,
        one_year_date=emp.one_year_date,
        vacation_days_total=emp.vacation_days_total or 0,
        vacation_days_used=emp.vacation_days_used or 0,
        sick_days_total=emp.sick_days_total or 10,
        sick_days_used=emp.sick_days_used or 0,
        family_leave_days_total=emp.family_leave_days_total or 3,
        family_leave_days_used=emp.family_leave_days_used or 0,
        nda_signed=emp.nda_signed or False,
        nda_signed_at=emp.nda_signed_at,
        contract_signed=emp.contract_signed or False,
        contract_signed_at=emp.contract_signed_at,
        is_active=emp.is_active if emp.is_active is not None else True,
        dismissed_at=emp.dismissed_at,
        dismissal_reason=emp.dismissal_reason,
        extra_data=emp.extra_data,
        created_at=emp.created_at,
        updated_at=emp.updated_at,
        user_name=emp.user.name if emp.user else None,
        user_email=emp.user.email if emp.user else None,
        department_name=emp.department.name if emp.department else None,
    )


# ─── Employee CRUD ───────────────────────────────────────────

async def sync_org_employees(org_id: int, db: AsyncSession) -> int:
    """Подтянуть всех участников организации («Управление») в HR/Factorial:
    создаёт запись Employee для каждого org-member, у кого её ещё нет. Идемпотентно."""
    member_ids = list((await db.execute(
        select(OrgMember.user_id).where(OrgMember.org_id == org_id)
    )).scalars().all())
    if not member_ids:
        return 0
    existing = set((await db.execute(
        select(Employee.user_id).where(Employee.user_id.in_(member_ids))
    )).scalars().all())
    created = 0
    for uid in member_ids:
        if uid not in existing:
            db.add(Employee(user_id=uid, org_id=org_id, is_active=True))
            existing.add(uid)
            created += 1
    if created:
        try:
            await db.commit()
        except Exception:
            await db.rollback()
    return created


async def sync_org_departments(org_id: int, db: AsyncSession) -> int:
    """Подтянуть распределение по отделам из Энцеладуса (departments/department_members)
    в Factorial: создать org_units по именам департаментов и проставить employees.org_unit_id
    тем, у кого он ещё НЕ задан. Идемпотентно; ручные назначения в Factorial не перетирает."""
    depts = (await db.execute(
        select(Department).where(Department.org_id == org_id, Department.is_active == True)  # noqa: E712
    )).scalars().all()
    if not depts:
        return 0
    units = (await db.execute(select(OrgUnit).where(OrgUnit.org_id == org_id))).scalars().all()
    unit_by_name = {u.name: u for u in units}
    dept_to_unit = {}
    created = 0
    for d in depts:
        u = unit_by_name.get(d.name)
        if not u:
            u = OrgUnit(org_id=org_id, name=d.name, parent_id=None, color=d.color, sort_order=0)
            db.add(u)
            await db.flush()
            unit_by_name[d.name] = u
            created += 1
        dept_to_unit[d.id] = u
    # вложенность отделов — только если у узла ещё нет родителя (не перетираем ручное)
    for d in depts:
        if d.parent_id and d.parent_id in dept_to_unit:
            u = dept_to_unit[d.id]
            parent_u = dept_to_unit[d.parent_id]
            if u.parent_id is None and parent_u.id != u.id:
                u.parent_id = parent_u.id
    # user_id -> department_id (первое членство)
    dept_ids = [d.id for d in depts]
    rows = (await db.execute(
        select(DepartmentMember.user_id, DepartmentMember.department_id)
        .where(DepartmentMember.department_id.in_(dept_ids))
    )).all()
    user_dept = {}
    for uid, did in rows:
        if uid not in user_dept:
            user_dept[uid] = did
    # проставить org_unit_id сотрудникам без отдела
    emps = (await db.execute(
        select(Employee).where(Employee.org_id == org_id, Employee.org_unit_id.is_(None))
    )).scalars().all()
    changed = 0
    for e in emps:
        did = user_dept.get(e.user_id)
        if did is None and e.department_id in dept_to_unit:
            did = e.department_id
        u = dept_to_unit.get(did) if did is not None else None
        if u:
            e.org_unit_id = u.id
            changed += 1
    if created or changed:
        try:
            await db.commit()
        except Exception:
            await db.rollback()
    return changed


@router.get("", response_model=List[EmployeeResponse])
async def list_employees(
    active_only: bool = Query(True),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List all employees (admin/HRD only)."""
    org = await get_user_org(current_user, db)
    if not org:
        raise HTTPException(status_code=403, detail="No organization")
    if not await _is_admin_or_owner(current_user, org.id, db):
        raise HTTPException(status_code=403, detail="Admin access required")

    try:
        await sync_org_employees(org.id, db)
        await sync_org_departments(org.id, db)
    except Exception:
        try:
            await db.rollback()
        except Exception:
            pass

    query = (
        select(Employee)
        .options(selectinload(Employee.user), selectinload(Employee.department))
        .where(Employee.org_id == org.id)
    )
    if active_only:
        query = query.where(Employee.is_active == True)
    query = query.order_by(Employee.created_at.desc())

    result = await db.execute(query)
    employees = list(result.scalars().all())

    # Update vacation days dynamically
    responses = []
    for emp in employees:
        emp.vacation_days_total = _calculate_vacation_days(emp.department_start_date)
        responses.append(_employee_to_response(emp))
    return responses


@router.post("/bulk-import")
async def bulk_import_employees(
    rows: List[BulkImportRow],
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    org = await get_user_org(current_user, db)
    if not org or not await _is_admin_or_owner(current_user, org.id, db):
        raise HTTPException(status_code=403, detail="Admin access required")
    updated, skipped, errors = 0, [], []
    for r in rows:
        email = (r.email or "").strip().lower()
        if not email:
            errors.append("пустой email в строке")
            continue
        user = (await db.execute(select(User).where(func.lower(User.email) == email))).scalar_one_or_none()
        emp = None
        if user:
            emp = (await db.execute(select(Employee).where(Employee.user_id == user.id, Employee.org_id == org.id))).scalar_one_or_none()
        if not emp:
            skipped.append(email)
            continue
        if r.position is not None:
            emp.position = r.position
        if r.phone is not None:
            emp.phone = r.phone
        if r.telegram_username is not None:
            emp.telegram_username = r.telegram_username
        if r.department_start_date:
            parsed = _parse_import_date(r.department_start_date)
            if parsed:
                emp.department_start_date = parsed
        patch = {}
        for k in ("last_name", "first_name", "middle_name", "address", "passport_number", "payment_method", "payment_details"):
            v = getattr(r, k)
            if v is not None and str(v).strip():
                patch[k] = str(v).strip()
        if patch.get('payment_method'):
            patch['payment_method'] = _normalize_pay_method(patch['payment_method'])
        emp.extra_data = _merge_extra(emp.extra_data, patch)
        updated += 1
    await db.commit()
    return {"updated": updated, "skipped": skipped, "errors": errors}


@router.get("/import-template")
async def import_template(
    filled: int = 0,
    id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    org = await get_user_org(current_user, db)
    if not org:
        raise HTTPException(status_code=403, detail="No organization")
    is_hr = await _is_admin_or_owner(current_user, org.id, db)

    rows = []
    with_example = False
    if id is not None:
        emp = (await db.execute(
            select(Employee).options(selectinload(Employee.user)).where(
                Employee.id == id, Employee.org_id == org.id
            )
        )).scalar_one_or_none()
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not is_hr and emp.user_id != current_user.id:
            raise HTTPException(status_code=403, detail="Access denied")
        rows = [_employee_to_template_row(emp)]
    else:
        if not is_hr:
            raise HTTPException(status_code=403, detail="Admin access required")
        if filled:
            emps = (await db.execute(
                select(Employee).options(selectinload(Employee.user)).where(Employee.org_id == org.id)
            )).scalars().all()
            rows = [_employee_to_template_row(e) for e in emps]
        else:
            with_example = True

    positions = [p for (p,) in (await db.execute(
        select(Employee.position).where(Employee.org_id == org.id, Employee.position.isnot(None))
    )).all()]
    data = _build_template_xlsx(rows=rows, positions=positions, with_example=with_example)
    fname = "employees.xlsx" if (filled or id) else "template_employees.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/me", response_model=EmployeeResponse)
async def get_my_profile(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get current user's employee profile."""
    result = await db.execute(
        select(Employee)
        .options(selectinload(Employee.user), selectinload(Employee.department))
        .where(Employee.user_id == current_user.id)
    )
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee profile not found")

    emp.vacation_days_total = _calculate_vacation_days(emp.department_start_date)
    return _employee_to_response(emp)


@router.put("/me", response_model=EmployeeResponse)
async def update_my_profile(
    data: EmployeeUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Сотрудник редактирует СВОЙ профиль (личный кабинет). Только своя запись, без админ-прав."""
    result = await db.execute(
        select(Employee)
        .options(selectinload(Employee.user), selectinload(Employee.department))
        .where(Employee.user_id == current_user.id)
    )
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee profile not found")

    # Self-service: сотрудник правит ТОЛЬКО безопасные поля. Привилегированные
    # (счётчики отпусков/больничных, подписание NDA/договора и их даты, дата
    # найма, отдел) меняет только HR через PUT /{employee_id}.
    SELF_EDITABLE = {"position", "phone", "telegram_username", "extra_data"}
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        if field not in SELF_EDITABLE:
            continue
        if field == "extra_data":
            emp.extra_data = _merge_extra(emp.extra_data, value)
        else:
            setattr(emp, field, value)

    emp.vacation_days_total = _calculate_vacation_days(emp.department_start_date)
    emp.updated_at = datetime.utcnow()

    await db.commit()
    await db.refresh(emp, attribute_names=["user", "department"])
    return _employee_to_response(emp)


@router.get("/leave-requests", response_model=List[LeaveRequestResponse])
async def list_all_leave_requests(
    status_filter: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List all pending leave requests (admin/HRD)."""
    org = await get_user_org(current_user, db)
    if not org:
        raise HTTPException(status_code=403, detail="No organization")
    if not await _is_admin_or_owner(current_user, org.id, db):
        raise HTTPException(status_code=403, detail="Admin access required")

    query = (
        select(LeaveRequest)
        .join(Employee, LeaveRequest.employee_id == Employee.id)
        .options(selectinload(LeaveRequest.employee).selectinload(Employee.user))
        .where(Employee.org_id == org.id)
    )
    if status_filter:
        query = query.where(LeaveRequest.status == status_filter)
    else:
        query = query.where(LeaveRequest.status == "pending")
    query = query.order_by(LeaveRequest.created_at.desc())

    result = await db.execute(query)
    requests = list(result.scalars().all())

    return [
        LeaveRequestResponse(
            id=lr.id,
            employee_id=lr.employee_id,
            type=lr.type,
            start_date=lr.start_date,
            end_date=lr.end_date,
            days=lr.days,
            reason=lr.reason,
            status=lr.status,
            approved_by=lr.approved_by,
            approved_at=lr.approved_at,
            created_at=lr.created_at,
            employee_name=lr.employee.user.name if lr.employee and lr.employee.user else None,
        )
        for lr in requests
    ]


@router.get("/reminders", response_model=List[ReminderItem])
async def get_reminders(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get upcoming reminders: probation ends within 14 days, 1 year within 14 days."""
    org = await get_user_org(current_user, db)
    if not org:
        raise HTTPException(status_code=403, detail="No organization")
    if not await _is_admin_or_owner(current_user, org.id, db):
        raise HTTPException(status_code=403, detail="Admin access required")

    now = datetime.utcnow()
    in_14_days = now + timedelta(days=14)

    result = await db.execute(
        select(Employee)
        .options(selectinload(Employee.user))
        .where(
            Employee.org_id == org.id,
            Employee.is_active == True,
            or_(
                and_(
                    Employee.probation_end_date != None,
                    Employee.probation_end_date >= now,
                    Employee.probation_end_date <= in_14_days,
                ),
                and_(
                    Employee.one_year_date != None,
                    Employee.one_year_date >= now,
                    Employee.one_year_date <= in_14_days,
                ),
            ),
        )
    )
    employees = list(result.scalars().all())

    reminders: List[ReminderItem] = []
    for emp in employees:
        name = emp.user.name if emp.user else f"Employee #{emp.id}"
        if emp.probation_end_date and now <= emp.probation_end_date <= in_14_days:
            reminders.append(ReminderItem(
                employee_id=emp.id,
                employee_name=name,
                type="probation_ending",
                date=emp.probation_end_date,
                days_remaining=(emp.probation_end_date - now).days,
            ))
        if emp.one_year_date and now <= emp.one_year_date <= in_14_days:
            reminders.append(ReminderItem(
                employee_id=emp.id,
                employee_name=name,
                type="one_year_anniversary",
                date=emp.one_year_date,
                days_remaining=(emp.one_year_date - now).days,
            ))

    reminders.sort(key=lambda r: r.days_remaining)
    return reminders


@router.get("/{employee_id}", response_model=EmployeeResponse)
async def get_employee(
    employee_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get employee detail."""
    result = await db.execute(
        select(Employee)
        .options(selectinload(Employee.user), selectinload(Employee.department))
        .where(Employee.id == employee_id)
    )
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    # Access check: admin or self
    org = await get_user_org(current_user, db)
    if emp.user_id != current_user.id:
        if not org or not await _is_admin_or_owner(current_user, emp.org_id, db):
            raise HTTPException(status_code=403, detail="Access denied")

    emp.vacation_days_total = _calculate_vacation_days(emp.department_start_date)
    return _employee_to_response(emp)


@router.post("", response_model=EmployeeResponse)
async def create_employee(
    data: EmployeeCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Create employee record (from practice list to staff)."""
    org = await get_user_org(current_user, db)
    if not org:
        raise HTTPException(status_code=403, detail="No organization")
    if not await _is_admin_or_owner(current_user, org.id, db):
        raise HTTPException(status_code=403, detail="Admin access required")

    # Check user exists
    user_result = await db.execute(select(User).where(User.id == data.user_id))
    target_user = user_result.scalar_one_or_none()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")

    # Check duplicate
    existing = await db.execute(select(Employee).where(Employee.user_id == data.user_id))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Employee record already exists for this user")

    emp = Employee(
        user_id=data.user_id,
        org_id=org.id,
        entity_id=data.entity_id,
        department_id=data.department_id,
        position=data.position,
        phone=data.phone,
        telegram_username=data.telegram_username,
        practice_start_date=data.practice_start_date,
        department_start_date=data.department_start_date,
        nda_signed=data.nda_signed,
        contract_signed=data.contract_signed,
        extra_data=data.extra_data or {},
    )
    _auto_calculate_dates(emp)
    emp.vacation_days_total = _calculate_vacation_days(emp.department_start_date)

    db.add(emp)
    await db.commit()
    await db.refresh(emp, attribute_names=["user", "department"])

    return _employee_to_response(emp)


@router.put("/{employee_id}", response_model=EmployeeResponse)
async def update_employee(
    employee_id: int,
    data: EmployeeUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update employee record."""
    org = await get_user_org(current_user, db)
    if not org:
        raise HTTPException(status_code=403, detail="No organization")

    result = await db.execute(
        select(Employee)
        .options(selectinload(Employee.user), selectinload(Employee.department))
        .where(Employee.id == employee_id, Employee.org_id == org.id)
    )
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    if not await _is_admin_or_owner(current_user, org.id, db):
        raise HTTPException(status_code=403, detail="Admin access required")

    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        if field == "extra_data":
            # Мёржим, чтобы частичный extra_data не затирал passport/documents и пр.
            emp.extra_data = _merge_extra(emp.extra_data, value)
        else:
            setattr(emp, field, value)

    # Recalculate dates if department_start_date changed
    if "department_start_date" in update_data:
        _auto_calculate_dates(emp)

    emp.vacation_days_total = _calculate_vacation_days(emp.department_start_date)
    emp.updated_at = datetime.utcnow()

    await db.commit()
    await db.refresh(emp, attribute_names=["user", "department"])
    return _employee_to_response(emp)


@router.delete("/{employee_id}")
async def dismiss_employee(
    employee_id: int,
    reason: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Dismiss (soft-delete) an employee."""
    org = await get_user_org(current_user, db)
    if not org:
        raise HTTPException(status_code=403, detail="No organization")
    if not await _is_admin_or_owner(current_user, org.id, db):
        raise HTTPException(status_code=403, detail="Admin access required")

    result = await db.execute(
        select(Employee).where(Employee.id == employee_id, Employee.org_id == org.id)
    )
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    emp.is_active = False
    emp.dismissed_at = datetime.utcnow()
    emp.dismissal_reason = reason
    await db.commit()

    return {"ok": True, "detail": "Employee dismissed"}


# ─── Leave balance & requests ────────────────────────────────

@router.get("/{employee_id}/leave-balance", response_model=LeaveBalanceResponse)
async def get_leave_balance(
    employee_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get leave balance for an employee."""
    result = await db.execute(select(Employee).where(Employee.id == employee_id))
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    # Access: self or admin
    if emp.user_id != current_user.id:
        org = await get_user_org(current_user, db)
        if not org or not await _is_admin_or_owner(current_user, emp.org_id, db):
            raise HTTPException(status_code=403, detail="Access denied")

    now = datetime.utcnow()
    vac_total = _calculate_vacation_days(emp.department_start_date, now=now)
    cstart = _cycle_start(emp.department_start_date, now=now)
    cend = (cstart + relativedelta(months=12)) if cstart else None

    # «Использовано» считаем по одобренным заявкам, чья ДАТА НАЧАЛА попадает в текущий
    # цикл (start_date >= cstart). Заявку относим к циклу по дате начала — по спеке.
    used = {"vacation": 0, "sick": 0, "family_leave": 0}
    if cstart:
        rows = (await db.execute(
            select(LeaveRequest.type, LeaveRequest.days).where(
                LeaveRequest.employee_id == emp.id,
                LeaveRequest.status == "approved",
                LeaveRequest.start_date >= cstart,
                LeaveRequest.start_date < cend,
            )
        )).all()
        for t, d in rows:
            if t in used:
                used[t] += (d or 0)

    sick_total = 7
    fl_total = 2
    return LeaveBalanceResponse(
        vacation_total=vac_total,
        vacation_used=used["vacation"],
        vacation_remaining=max(0, vac_total - used["vacation"]),
        sick_total=sick_total,
        sick_used=used["sick"],
        sick_remaining=max(0, sick_total - used["sick"]),
        family_leave_total=fl_total,
        family_leave_used=used["family_leave"],
        family_leave_remaining=max(0, fl_total - used["family_leave"]),
        cycle_start=cstart,
        cycle_end=cend,
    )


@router.post("/{employee_id}/leave-request", response_model=LeaveRequestResponse)
async def create_leave_request(
    employee_id: int,
    data: LeaveRequestCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Request leave (employee self-service)."""
    result = await db.execute(
        select(Employee)
        .options(selectinload(Employee.user))
        .where(Employee.id == employee_id)
    )
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    # Only self or admin can create
    if emp.user_id != current_user.id:
        org = await get_user_org(current_user, db)
        if not org or not await _is_admin_or_owner(current_user, emp.org_id, db):
            raise HTTPException(status_code=403, detail="Access denied")

    # Validate type
    if data.type not in ("vacation", "sick", "family_leave", "bereavement"):
        raise HTTPException(status_code=400, detail="Invalid leave type")

    lr = LeaveRequest(
        employee_id=employee_id,
        type=data.type,
        start_date=data.start_date,
        end_date=data.end_date,
        days=data.days,
        reason=data.reason,
        status="pending",
    )
    db.add(lr)
    await db.commit()
    await db.refresh(lr)

    return LeaveRequestResponse(
        id=lr.id,
        employee_id=lr.employee_id,
        type=lr.type,
        start_date=lr.start_date,
        end_date=lr.end_date,
        days=lr.days,
        reason=lr.reason,
        status=lr.status,
        approved_by=lr.approved_by,
        approved_at=lr.approved_at,
        created_at=lr.created_at,
        employee_name=emp.user.name if emp.user else None,
    )


@router.put("/leave-requests/{request_id}/approve")
async def approve_leave_request(
    request_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Approve a leave request (admin/HRD)."""
    org = await get_user_org(current_user, db)
    if not org:
        raise HTTPException(status_code=403, detail="No organization")
    if not await _is_admin_or_owner(current_user, org.id, db):
        raise HTTPException(status_code=403, detail="Admin access required")

    result = await db.execute(
        select(LeaveRequest)
        .join(Employee)
        .where(LeaveRequest.id == request_id, Employee.org_id == org.id)
    )
    lr = result.scalar_one_or_none()
    if not lr:
        raise HTTPException(status_code=404, detail="Leave request not found")
    if lr.status != "pending":
        raise HTTPException(status_code=400, detail="Request is not pending")

    # Сотрудник нужен и для проверки лимита, и для обновления счётчиков (один запрос).
    emp_result = await db.execute(select(Employee).where(Employee.id == lr.employee_id))
    emp = emp_result.scalar_one_or_none()

    # Проверка остатка: одобрять отпуск можно, пока в текущем рабочем году есть
    # доступные накопленные дни (дробление по числу заявок НЕ ограничено).
    if lr.type == "vacation" and emp:
        now = datetime.utcnow()
        cstart = _cycle_start(emp.department_start_date, now=now)
        if cstart:
            cend = cstart + relativedelta(months=12)
            accrued = _calculate_vacation_days(emp.department_start_date, now=now)
            used_days = (await db.execute(
                select(func.coalesce(func.sum(LeaveRequest.days), 0)).where(
                    LeaveRequest.employee_id == lr.employee_id,
                    LeaveRequest.type == "vacation",
                    LeaveRequest.status == "approved",
                    LeaveRequest.start_date >= cstart,
                    LeaveRequest.start_date < cend,
                )
            )).scalar() or 0
            available = max(0, accrued - used_days)
            if lr.days > available:
                raise HTTPException(
                    status_code=400,
                    detail=f"Недостаточно дней отпуска: доступно {available}",
                )

    lr.status = "approved"
    lr.approved_by = current_user.id
    lr.approved_at = datetime.utcnow()

    # Обновляем legacy-счётчики сотрудника (их читает старый HR-портал).
    if emp:
        if lr.type == "vacation":
            emp.vacation_days_used = (emp.vacation_days_used or 0) + lr.days
        elif lr.type == "sick":
            emp.sick_days_used = (emp.sick_days_used or 0) + lr.days
        elif lr.type == "family_leave":
            emp.family_leave_days_used = (emp.family_leave_days_used or 0) + lr.days

    await db.commit()
    return {"ok": True, "detail": "Leave request approved"}


@router.put("/leave-requests/{request_id}/reject")
async def reject_leave_request(
    request_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Reject a leave request (admin/HRD)."""
    org = await get_user_org(current_user, db)
    if not org:
        raise HTTPException(status_code=403, detail="No organization")
    if not await _is_admin_or_owner(current_user, org.id, db):
        raise HTTPException(status_code=403, detail="Admin access required")

    result = await db.execute(
        select(LeaveRequest)
        .join(Employee)
        .where(LeaveRequest.id == request_id, Employee.org_id == org.id)
    )
    lr = result.scalar_one_or_none()
    if not lr:
        raise HTTPException(status_code=404, detail="Leave request not found")
    if lr.status != "pending":
        raise HTTPException(status_code=400, detail="Request is not pending")

    lr.status = "rejected"
    lr.approved_by = current_user.id
    lr.approved_at = datetime.utcnow()
    await db.commit()

    return {"ok": True, "detail": "Leave request rejected"}


# ─── Secure passport upload (employee self-service, encrypted at rest) ───

class PassportUpload(BaseModel):
    filename: str
    content_type: str
    data_base64: str  # raw file bytes, base64-encoded (no data: prefix)


def _passport_fernet():
    import base64 as _b64
    import hashlib
    from cryptography.fernet import Fernet
    from api.config import settings
    key = _b64.urlsafe_b64encode(hashlib.sha256(settings.jwt_secret.encode()).digest())
    return Fernet(key)


def _passport_path(emp_id: int):
    from pathlib import Path
    d = Path("uploads") / "passports"
    d.mkdir(parents=True, exist_ok=True)
    return d / f"{emp_id}.enc"


def _employee_docs_dir():
    from pathlib import Path
    d = Path("uploads") / "employee_docs"
    d.mkdir(parents=True, exist_ok=True)
    return d


async def _can_view_employee(current_user: User, emp: Employee, db: AsyncSession) -> bool:
    """True если current_user может смотреть карточку emp: сам сотрудник, admin/owner, HR, или его руководитель.
    ВАЖНО: сверь набор с существующим passport-эндпоинтом get_employee_passport и приведи к тому же."""
    if emp.user_id == current_user.id:
        return True
    if await _is_admin_or_owner(current_user, emp.org_id, db):
        return True
    org = await get_user_org(current_user, db)
    if org:
        member = (await db.execute(
            select(OrgMember).where(OrgMember.user_id == current_user.id, OrgMember.org_id == emp.org_id)
        )).scalar_one_or_none()
        if member and member.role == OrgRole.hr:
            return True
    if emp.manager_id:
        mgr = (await db.execute(select(Employee).where(Employee.id == emp.manager_id))).scalar_one_or_none()
        if mgr and mgr.user_id == current_user.id:
            return True
    return False


@router.post("/me/passport")
async def upload_my_passport(
    payload: PassportUpload,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Upload own passport scan. Stored encrypted on disk; readable only by the employee."""
    import base64 as _b64
    result = await db.execute(select(Employee).where(Employee.user_id == current_user.id))
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee profile not found")
    try:
        raw = _b64.b64decode(payload.data_base64)
    except Exception:
        raise HTTPException(status_code=400, detail="Некорректные данные файла")
    if len(raw) > 8 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Файл слишком большой (макс 8 МБ)")
    enc = _passport_fernet().encrypt(raw)
    _passport_path(emp.id).write_bytes(enc)
    # Дублируем в БД (bytea): диск на проде эфемерный, иначе паспорт теряется.
    emp.passport_data = enc
    extra = dict(emp.extra_data or {})
    extra["passport"] = {
        "filename": payload.filename,
        "content_type": payload.content_type,
        "uploaded_at": datetime.utcnow().isoformat(),
    }
    emp.extra_data = extra
    await db.commit()
    return {"ok": True, "filename": payload.filename}


@router.get("/me/passport")
async def get_my_passport(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Download own passport scan (decrypted). Only the employee can access their own file."""
    import base64 as _b64
    result = await db.execute(select(Employee).where(Employee.user_id == current_user.id))
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee profile not found")
    meta = (emp.extra_data or {}).get("passport")
    # Сначала зашифрованные байты из БД (переживают редеплой), затем диск (legacy).
    enc = emp.passport_data
    if enc is None:
        path = _passport_path(emp.id)
        enc = path.read_bytes() if path.exists() else None
    if not meta or enc is None:
        raise HTTPException(status_code=404, detail="Паспорт не загружен")
    raw = _passport_fernet().decrypt(enc)
    return {
        "filename": meta.get("filename"),
        "content_type": meta.get("content_type"),
        "data_base64": _b64.b64encode(raw).decode(),
    }


@router.get("/{employee_id}/passport")
async def get_employee_passport(
    employee_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Скан паспорта сотрудника. Доступ: сам сотрудник, HR (owner/admin/hr) или его руководитель."""
    import base64 as _b64
    emp = (await db.execute(select(Employee).where(Employee.id == employee_id))).scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    allowed = emp.user_id == current_user.id  # сам сотрудник
    if not allowed and await _is_admin_or_owner(current_user, emp.org_id, db):
        allowed = True  # superadmin / owner / admin (HR-админ)
    if not allowed:
        om = (await db.execute(
            select(OrgMember).where(OrgMember.user_id == current_user.id, OrgMember.org_id == emp.org_id)
        )).scalar_one_or_none()
        if om and om.role == OrgRole.hr:
            allowed = True  # HR-рекрутер
    if not allowed and emp.manager_id:
        mgr = (await db.execute(
            select(Employee).where(Employee.id == emp.manager_id, Employee.org_id == emp.org_id)
        )).scalar_one_or_none()
        if mgr and mgr.user_id == current_user.id:
            allowed = True  # непосредственный руководитель
    if not allowed:
        raise HTTPException(status_code=403, detail="Нет доступа к паспорту этого сотрудника")

    meta = (emp.extra_data or {}).get("passport")
    # Сначала зашифрованные байты из БД (переживают редеплой), затем диск (legacy).
    enc = emp.passport_data
    if enc is None:
        path = _passport_path(emp.id)
        enc = path.read_bytes() if path.exists() else None
    if not meta or enc is None:
        raise HTTPException(status_code=404, detail="Паспорт не загружен")
    raw = _passport_fernet().decrypt(enc)
    return {
        "filename": meta.get("filename"),
        "content_type": meta.get("content_type"),
        "data_base64": _b64.b64encode(raw).decode(),
    }


@router.get("/{employee_id}/documents")
async def list_employee_documents(employee_id: int, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    emp = (await db.execute(select(Employee).where(Employee.id == employee_id))).scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    if not await _can_view_employee(current_user, emp, db):
        raise HTTPException(status_code=403, detail="Access denied")
    docs = (await db.execute(
        select(EmployeeDocument).where(EmployeeDocument.employee_id == employee_id).order_by(EmployeeDocument.uploaded_at.desc())
    )).scalars().all()
    return [{"id": d.id, "filename": d.filename, "content_type": d.content_type, "size": d.size, "uploaded_at": d.uploaded_at} for d in docs]


@router.post("/{employee_id}/documents")
async def upload_employee_document(employee_id: int, request: Request, file: UploadFile = File(...), db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    emp = (await db.execute(select(Employee).where(Employee.id == employee_id))).scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    # загрузка: сам сотрудник или HR/admin/owner
    allowed = emp.user_id == current_user.id or await _is_admin_or_owner(current_user, emp.org_id, db)
    if not allowed:
        org = await get_user_org(current_user, db)
        member = (await db.execute(select(OrgMember).where(OrgMember.user_id == current_user.id, OrgMember.org_id == emp.org_id))).scalar_one_or_none() if org else None
        allowed = bool(member and member.role == OrgRole.hr)
    if not allowed:
        raise HTTPException(status_code=403, detail="Access denied")
    clen = request.headers.get("content-length")
    if clen and clen.isdigit() and int(clen) > 12 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Файл больше 10 МБ")
    raw = await file.read()
    if len(raw) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Файл больше 10 МБ")
    import uuid
    enc = _passport_fernet().encrypt(raw)
    fpath = _employee_docs_dir() / f"{employee_id}_{uuid.uuid4().hex}.enc"
    fpath.write_bytes(enc)
    # file_data (bytea) переживает редеплой; диск остаётся legacy-fallback.
    doc = EmployeeDocument(employee_id=employee_id, filename=file.filename, content_type=file.content_type, size=len(raw), path=str(fpath), file_data=enc, uploaded_by=current_user.id)
    db.add(doc)
    await db.commit()
    await db.refresh(doc)
    return {"id": doc.id, "filename": doc.filename, "content_type": doc.content_type, "size": doc.size, "uploaded_at": doc.uploaded_at}


@router.get("/{employee_id}/documents/{doc_id}")
async def download_employee_document(employee_id: int, doc_id: int, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    emp = (await db.execute(select(Employee).where(Employee.id == employee_id))).scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    if not await _can_view_employee(current_user, emp, db):
        raise HTTPException(status_code=403, detail="Access denied")
    doc = (await db.execute(select(EmployeeDocument).where(EmployeeDocument.id == doc_id, EmployeeDocument.employee_id == employee_id))).scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    import base64
    # file_data (bytea) первично — переживает редеплой; диск как legacy-fallback.
    enc = doc.file_data
    if enc is None:
        from pathlib import Path
        p = Path(doc.path)
        enc = p.read_bytes() if p.exists() else None
    if enc is None:
        raise HTTPException(status_code=410, detail="Файл недоступен (потерян при обновлении сервера)")
    data = _passport_fernet().decrypt(enc)
    return {"filename": doc.filename, "content_type": doc.content_type, "data_base64": base64.b64encode(data).decode()}


@router.delete("/{employee_id}/documents/{doc_id}")
async def delete_employee_document(employee_id: int, doc_id: int, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    emp = (await db.execute(select(Employee).where(Employee.id == employee_id))).scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    doc = (await db.execute(select(EmployeeDocument).where(EmployeeDocument.id == doc_id, EmployeeDocument.employee_id == employee_id))).scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    allowed = doc.uploaded_by == current_user.id or await _is_admin_or_owner(current_user, emp.org_id, db)
    if not allowed:
        org = await get_user_org(current_user, db)
        member = (await db.execute(select(OrgMember).where(OrgMember.user_id == current_user.id, OrgMember.org_id == emp.org_id))).scalar_one_or_none() if org else None
        allowed = bool(member and member.role == OrgRole.hr)
    if not allowed:
        raise HTTPException(status_code=403, detail="Access denied")
    from pathlib import Path
    Path(doc.path).unlink(missing_ok=True)
    await db.delete(doc)
    await db.commit()
    return {"ok": True}
